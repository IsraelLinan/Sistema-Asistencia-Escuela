import customtkinter as ctk
from tkinter import messagebox, filedialog, ttk
from tkcalendar import DateEntry
from colegio_lib import db_pool, COLORS
from datetime import datetime
import csv
from PIL import Image
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import openpyxl

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")


class WeeklyReportsModule(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color=COLORS["bg"])
        self.parent = parent
        if isinstance(parent, ctk.CTk) or isinstance(parent, ctk.CTkToplevel):
            self.parent.title("Módulo Reporte de Asistencias")
        self.report_data = []
        self._build_ui()

    def _build_ui(self):
        # ── Encabezado ───────────────────────────────────────────────────────
        header = ctk.CTkFrame(self, fg_color=COLORS["bg2"], corner_radius=14,
                              border_width=1, border_color=COLORS["border"])
        header.pack(fill="x", padx=20, pady=(20, 10))

        try:
            logo_img = ctk.CTkImage(
                light_image=Image.open("logo_colegio.png"),
                dark_image=Image.open("logo_colegio.png"),
                size=(70, 70)
            )
            ctk.CTkLabel(header, image=logo_img, text="").pack(side="left", padx=18, pady=12)
        except Exception:
            ctk.CTkLabel(header, text="📋", font=("Segoe UI", 36)).pack(side="left", padx=18, pady=12)

        title_box = ctk.CTkFrame(header, fg_color="transparent")
        title_box.pack(side="left", pady=12)
        ctk.CTkLabel(title_box, text="Reportes de Asistencia",
                     font=ctk.CTkFont("Segoe UI", 18, "bold"),
                     text_color=COLORS["text"]).pack(anchor="w")
        ctk.CTkLabel(title_box, text="Consulte y exporte los registros por rango de fechas",
                     font=ctk.CTkFont("Segoe UI", 11),
                     text_color=COLORS["text_muted"]).pack(anchor="w")

        # ── Controles de filtro ──────────────────────────────────────────────
        filters = ctk.CTkFrame(self, fg_color=COLORS["bg2"], corner_radius=14,
                               border_width=1, border_color=COLORS["border"])
        filters.pack(fill="x", padx=20, pady=6)

        row1 = ctk.CTkFrame(filters, fg_color="transparent")
        row1.pack(fill="x", padx=20, pady=(16, 8))

        # Fecha inicio
        ctk.CTkLabel(row1, text="Fecha Inicio", font=ctk.CTkFont("Segoe UI", 11, "bold"),
                     text_color=COLORS["text_muted"], width=90, anchor="w").pack(side="left")
        self.start_date_entry = DateEntry(row1, width=12, background=COLORS["accent"],
                                         foreground="white", borderwidth=0,
                                         headersbackground=COLORS["bg2"],
                                         normalbackground=COLORS["bg3"],
                                         weekendbackground=COLORS["bg3"],
                                         othermonthbackground=COLORS["bg"],
                                         headersforeground=COLORS["text"],
                                         normalforeground=COLORS["text"],
                                         font=("Segoe UI", 10))
        self.start_date_entry.pack(side="left", padx=(4, 24))

        # Fecha fin
        ctk.CTkLabel(row1, text="Fecha Fin", font=ctk.CTkFont("Segoe UI", 11, "bold"),
                     text_color=COLORS["text_muted"], width=70, anchor="w").pack(side="left")
        self.end_date_entry = DateEntry(row1, width=12, background=COLORS["accent"],
                                       foreground="white", borderwidth=0,
                                       headersbackground=COLORS["bg2"],
                                       normalbackground=COLORS["bg3"],
                                       weekendbackground=COLORS["bg3"],
                                       othermonthbackground=COLORS["bg"],
                                       headersforeground=COLORS["text"],
                                       normalforeground=COLORS["text"],
                                       font=("Segoe UI", 10))
        self.end_date_entry.pack(side="left", padx=(4, 24))

        # Tipo
        ctk.CTkLabel(row1, text="Tipo", font=ctk.CTkFont("Segoe UI", 11, "bold"),
                     text_color=COLORS["text_muted"], width=40, anchor="w").pack(side="left")
        self.type_menu = ctk.CTkOptionMenu(
            row1, values=["Todos", "Estudiantes", "Docentes"],
            width=140, height=34, corner_radius=8,
            fg_color=COLORS["bg3"], button_color=COLORS["accent"],
            button_hover_color=COLORS["accent_hover"],
            text_color=COLORS["text"], font=ctk.CTkFont("Segoe UI", 11)
        )
        self.type_menu.set("Todos")
        self.type_menu.pack(side="left", padx=(4, 0))

        # Botón cargar
        ctk.CTkButton(row1, text="⟳  Cargar", width=110, height=34,
                      corner_radius=8, font=ctk.CTkFont("Segoe UI", 11, "bold"),
                      fg_color=COLORS["accent"], hover_color=COLORS["accent_hover"],
                      command=self.load_report_data).pack(side="right")

        # Buscador
        row2 = ctk.CTkFrame(filters, fg_color="transparent")
        row2.pack(fill="x", padx=20, pady=(0, 16))

        ctk.CTkLabel(row2, text="🔍", font=("Segoe UI", 14),
                     text_color=COLORS["text_muted"]).pack(side="left", padx=(0, 6))
        self.search_entry = ctk.CTkEntry(
            row2, placeholder_text="Buscar por nombre o tipo...",
            height=34, corner_radius=8,
            fg_color=COLORS["bg3"], border_color=COLORS["border"],
            text_color=COLORS["text"], font=ctk.CTkFont("Segoe UI", 11)
        )
        self.search_entry.pack(side="left", fill="x", expand=True)
        self.search_entry.bind("<KeyRelease>", self.filter_report)

        # ── Tabla (Treeview nativo con estilo oscuro) ─────────────────────────
        tree_frame = ctk.CTkFrame(self, fg_color=COLORS["bg2"], corner_radius=14,
                                  border_width=1, border_color=COLORS["border"])
        tree_frame.pack(fill="both", expand=True, padx=20, pady=6)

        # Estilo para el Treeview
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Dark.Treeview",
                        background=COLORS["bg2"],
                        foreground=COLORS["text"],
                        rowheight=32,
                        fieldbackground=COLORS["bg2"],
                        borderwidth=0,
                        font=("Segoe UI", 10))
        style.configure("Dark.Treeview.Heading",
                        background=COLORS["bg3"],
                        foreground=COLORS["text_muted"],
                        font=("Segoe UI", 10, "bold"),
                        relief="flat",
                        borderwidth=0)
        style.map("Dark.Treeview",
                  background=[("selected", COLORS["accent"])],
                  foreground=[("selected", "#ffffff")])
        style.map("Dark.Treeview.Heading",
                  background=[("active", COLORS["border"])])

        columns = ("Tipo", "Nombre", "Hora Ingreso", "Hora Salida")
        self.report_tree = ttk.Treeview(tree_frame, columns=columns,
                                        show="headings", style="Dark.Treeview",
                                        selectmode="browse")

        for col, w in zip(columns, [110, 280, 170, 170]):
            self.report_tree.heading(col, text=col)
            self.report_tree.column(col, width=w, anchor="w" if col == "Nombre" else "center")

        scrollbar = ctk.CTkScrollbar(tree_frame, command=self.report_tree.yview,
                                     fg_color=COLORS["bg2"],
                                     button_color=COLORS["border"],
                                     button_hover_color=COLORS["accent"])
        self.report_tree.configure(yscrollcommand=scrollbar.set)

        self.report_tree.pack(side="left", fill="both", expand=True, padx=(8, 0), pady=8)
        scrollbar.pack(side="right", fill="y", pady=8, padx=(0, 6))

        # Alternar colores de filas
        self.report_tree.tag_configure("odd", background=COLORS["bg2"])
        self.report_tree.tag_configure("even", background=COLORS["bg3"])

        # ── Barra inferior: estadísticas + exportar ──────────────────────────
        bottom = ctk.CTkFrame(self, fg_color=COLORS["bg2"], corner_radius=14,
                              border_width=1, border_color=COLORS["border"])
        bottom.pack(fill="x", padx=20, pady=(6, 20))

        self.stats_label = ctk.CTkLabel(bottom, text="Sin datos cargados",
                                         font=ctk.CTkFont("Segoe UI", 11),
                                         text_color=COLORS["text_muted"])
        self.stats_label.pack(side="left", padx=20, pady=12)

        export_row = ctk.CTkFrame(bottom, fg_color="transparent")
        export_row.pack(side="right", padx=16, pady=10)

        btn_style = dict(height=36, corner_radius=8,
                         font=ctk.CTkFont("Segoe UI", 11),
                         fg_color=COLORS["bg3"], hover_color=COLORS["border"],
                         border_width=1, border_color=COLORS["border"])

        ctk.CTkButton(export_row, text="CSV", width=80, **btn_style,
                      command=self.export_to_csv).pack(side="left", padx=4)
        ctk.CTkButton(export_row, text="PDF", width=80, **btn_style,
                      command=self.export_to_pdf).pack(side="left", padx=4)
        ctk.CTkButton(export_row, text="Excel", width=80,
                      height=36, corner_radius=8,
                      font=ctk.CTkFont("Segoe UI", 11),
                      fg_color=COLORS["accent"], hover_color=COLORS["accent_hover"],
                      command=self.export_to_xlsx).pack(side="left", padx=4)

    # ── Lógica ───────────────────────────────────────────────────────────────

    def load_report_data(self):
        start_date = self.start_date_entry.get_date()
        end_date = self.end_date_entry.get_date()
        person_type = self.type_menu.get()
        self.report_data = []
        conn = None
        try:
            conn = db_pool.get_conn()
            cur = conn.cursor()

            if person_type in ["Todos", "Estudiantes"]:
                cur.execute(
                    "SELECT 'Estudiante', e.nombre, i.hora_ingreso, i.hora_salida FROM ingresos_estudiantes i JOIN estudiantes e ON i.estudiante_id = e.id WHERE DATE(i.hora_ingreso) BETWEEN %s AND %s ORDER BY i.hora_ingreso",
                    (start_date, end_date)
                )
                for rec in cur.fetchall():
                    hora_salida = rec[3].strftime('%Y-%m-%d %H:%M:%S') if rec[3] else 'No registrada'
                    self.report_data.append({'tipo': rec[0], 'nombre': rec[1],
                                             'hora_ingreso': rec[2].strftime('%Y-%m-%d %H:%M:%S'),
                                             'hora_salida': hora_salida})

            if person_type in ["Todos", "Docentes"]:
                cur.execute(
                    "SELECT 'Docente', d.nombre, i.hora_ingreso, i.hora_salida FROM ingresos_docentes i JOIN docentes d ON i.docente_id = d.id WHERE DATE(i.hora_ingreso) BETWEEN %s AND %s ORDER BY i.hora_ingreso",
                    (start_date, end_date)
                )
                for rec in cur.fetchall():
                    hora_salida = rec[3].strftime('%Y-%m-%d %H:%M:%S') if rec[3] else 'No registrada'
                    self.report_data.append({'tipo': rec[0], 'nombre': rec[1],
                                             'hora_ingreso': rec[2].strftime('%Y-%m-%d %H:%M:%S'),
                                             'hora_salida': hora_salida})

            cur.close()
            self._refresh_tree(self.report_data)
            total = len(self.report_data)
            est = sum(1 for r in self.report_data if r['tipo'] == 'Estudiante')
            doc = total - est
            self.stats_label.configure(
                text=f"Total: {total} registros  •  Estudiantes: {est}  •  Docentes: {doc}"
            )

        except Exception as e:
            messagebox.showerror("Error", f"Error al cargar el reporte: {e}")
        finally:
            if conn:
                db_pool.put_conn(conn)

    def _refresh_tree(self, data):
        for item in self.report_tree.get_children():
            self.report_tree.delete(item)
        for i, record in enumerate(data):
            tag = "even" if i % 2 == 0 else "odd"
            self.report_tree.insert("", "end", tags=(tag,), values=(
                record['tipo'], record['nombre'],
                record['hora_ingreso'], record['hora_salida']
            ))

    def filter_report(self, event=None):
        query = self.search_entry.get().lower()
        filtered = [r for r in self.report_data
                    if query in r['nombre'].lower() or query in r['tipo'].lower()]
        self._refresh_tree(filtered)

    def refresh_treeview(self):
        self._refresh_tree(self.report_data)

    # ── Exportaciones ────────────────────────────────────────────────────────

    def export_to_csv(self):
        if not self.report_data:
            messagebox.showwarning("Atención", "No hay datos para exportar.")
            return
        default_name = f"Reporte_Asistencia_{datetime.now().strftime('%Y-%m-%d')}.csv"
        filepath = filedialog.asksaveasfilename(defaultextension=".csv",
                                                filetypes=[("CSV files", "*.csv")],
                                                initialfile=default_name)
        if not filepath:
            return
        try:
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=['tipo', 'nombre', 'hora_ingreso', 'hora_salida'])
                writer.writeheader()
                writer.writerows(self.report_data)
            messagebox.showinfo("Éxito", f"CSV guardado en:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Error CSV", str(e))

    def export_to_pdf(self):
        if not self.report_data:
            messagebox.showwarning("Atención", "No hay datos para exportar.")
            return
        default_name = f"Reporte_Asistencia_{datetime.now().strftime('%Y-%m-%d')}.pdf"
        filepath = filedialog.asksaveasfilename(defaultextension=".pdf",
                                                filetypes=[("PDF files", "*.pdf")],
                                                initialfile=default_name)
        if not filepath:
            return
        c = canvas.Canvas(filepath, pagesize=letter)
        c.setFont("Helvetica-Bold", 14)
        c.drawString(30, 760, "Reporte de Asistencia")
        c.setFont("Helvetica", 9)
        c.drawString(30, 745, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        c.line(30, 740, 580, 740)

        y = 720
        c.setFont("Helvetica-Bold", 10)
        for x, label in [(40, "Tipo"), (140, "Nombre"), (340, "Hora Ingreso"), (470, "Hora Salida")]:
            c.drawString(x, y, label)
        y -= 18
        c.setFont("Helvetica", 9)

        for record in self.report_data:
            if y < 50:
                c.showPage()
                y = 750
                c.setFont("Helvetica-Bold", 10)
                for x, label in [(40, "Tipo"), (140, "Nombre"), (340, "Hora Ingreso"), (470, "Hora Salida")]:
                    c.drawString(x, y, label)
                y -= 18
                c.setFont("Helvetica", 9)
            c.drawString(40, y, record['tipo'])
            c.drawString(140, y, record['nombre'])
            c.drawString(340, y, record['hora_ingreso'])
            c.drawString(470, y, record['hora_salida'])
            y -= 15

        c.save()
        messagebox.showinfo("Éxito", f"PDF guardado en:\n{filepath}")

    def export_to_xlsx(self):
        if not self.report_data:
            messagebox.showwarning("Atención", "No hay datos para exportar.")
            return
        default_name = f"Reporte_Asistencia_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                filetypes=[("Excel files", "*.xlsx")],
                                                initialfile=default_name)
        if not filepath:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Asistencia"
        headers = ["Tipo", "Nombre", "Hora Ingreso", "Hora Salida"]
        ws.append(headers)
        bold_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        fill = openpyxl.styles.PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        for i, cell in enumerate(ws[1], 1):
            cell.font = bold_font
            cell.fill = fill
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 28
        for record in self.report_data:
            ws.append([record['tipo'], record['nombre'],
                       record['hora_ingreso'], record['hora_salida']])
        wb.save(filepath)
        messagebox.showinfo("Éxito", f"Excel guardado en:\n{filepath}")


if __name__ == "__main__":
    root = ctk.CTk()
    root.configure(fg_color=COLORS["bg"])
    root.geometry("860x640")
    app = WeeklyReportsModule(root)
    app.pack(fill="both", expand=True, padx=10, pady=10)
    root.mainloop()













